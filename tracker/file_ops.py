import csv
import json
import os
import shutil
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CSV_PATH = os.path.join(PROJECT_ROOT, 'wydatki.csv')
RECURRING_PATH = os.path.join(PROJECT_ROOT, 'recurring.csv')
BACKUP_DIR = os.path.join(PROJECT_ROOT, 'backups')
BACKUP_EMERGENCY_DIR = os.path.join(PROJECT_ROOT, 'emergency backups')
BUDGET_PATH = os.path.join(PROJECT_ROOT, 'budget.csv')
EXPORTS_DIR = os.path.join(PROJECT_ROOT, 'exports')

def file_verification_main():
    if not os.path.exists(CSV_PATH):
        with open(CSV_PATH, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria'])
        print(f"Utworzono nowy plik: {CSV_PATH}")
    else:
        with open(CSV_PATH, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)
        expected_columns = ['ID', 'Data', 'Opis', 'Kwota', 'Kategoria']
        if header != expected_columns:
            create_emergency_backup(CSV_PATH)
            with open(CSV_PATH, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria'])

def get_all_expenses_main():
    with open(CSV_PATH, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        return list(reader)

def write_all_expenses_main(all_rows):
    with open(CSV_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria'])
        writer.writerows(all_rows)

def add_new_expense_main(expense_id, date, description, amount, category):
    with open(CSV_PATH, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([expense_id, date, description, amount, category])

def create_backup(csv_file_path):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")

    basename = os.path.basename(csv_file_path)

    if basename == 'wydatki.csv':
        backup_filename = f"wydatki_backup_{timestamp}.csv"
    elif basename == 'recurring.csv':
        backup_filename = f"recurring_backup_{timestamp}.csv"
    elif basename == 'budget.csv':
        backup_filename = f"budget_backup_{timestamp}.csv"
    else:
        raise ValueError(f"Unknown CSV file: {basename}")
    backup_fullpath = os.path.join(BACKUP_DIR, backup_filename)
    shutil.copyfile(csv_file_path, backup_fullpath)
    # print(f"Utworzono kopie zapasową: {backup_fullpath}")
    delete_old_backups(basename)

def delete_old_backups(file_type):
    prefix = f"{file_type.replace('.csv','')}_backup"
    backups = [f for f in os.listdir(BACKUP_DIR) if f.startswith(prefix) and f.endswith(".csv")]
    if len(backups) > 20:
        backups.sort()
        num_to_delete = len(backups) - 20
        for old_backup in backups[:num_to_delete]:
            backup_for_removal = os.path.join(BACKUP_DIR, old_backup)
            os.remove(backup_for_removal)

def create_emergency_backup(csv_file_path):
    if not os.path.exists(csv_file_path):
        return
    os.makedirs(BACKUP_EMERGENCY_DIR, exist_ok=True)
    now = datetime.now()
    timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    basename = os.path.basename(csv_file_path)
    name, ext = os.path.splitext(basename)
    backup_filename = f"{name}_emergency_backup_{timestamp}{ext}"
    backup_fullpath = os.path.join(BACKUP_EMERGENCY_DIR, backup_filename)
    shutil.copyfile(csv_file_path, backup_fullpath)
    print(f"Plik {basename} już istnieje, ale w innym formacie, utworzono jego kopię {backup_fullpath} oraz nadpisano do poprawnego formatu")

def file_verification_recurring():
    if not os.path.exists(RECURRING_PATH):
        with open(RECURRING_PATH, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria', 'Częstotliwość'])
        print(f"Utworzono nowy plik: {RECURRING_PATH}")
    else:
        with open(RECURRING_PATH, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)
        expected_columns = ['ID', 'Data', 'Opis', 'Kwota', 'Kategoria', 'Częstotliwość']
        if header != expected_columns:
            create_emergency_backup(RECURRING_PATH)
            with open(RECURRING_PATH, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria', 'Częstotliwość'])

def load_recurring_expenses():
    if not os.path.exists(RECURRING_PATH):
        return []
    with open(RECURRING_PATH, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        return list(reader)

def add_new_recurring_expense(expense_id, date, description, amount, category, frequency):
    with open(RECURRING_PATH, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([expense_id, date, description, amount, category, frequency])

def write_all_recurring_expenses(all_rows):
    with open(RECURRING_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['ID', 'Data', 'Opis', 'Kwota', 'Kategoria', 'Częstotliwość'])
        writer.writerows(all_rows)

def file_verification_budget():
    if not os.path.exists(BUDGET_PATH):
        with open(BUDGET_PATH, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['ID', 'Rok', 'Miesiąc', 'Kwota', 'Status'])
        print(f"Utworzono nowy plik: {BUDGET_PATH}")
    else:
        with open(BUDGET_PATH, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            header = next(reader)
        expected_columns = ['ID', 'Rok', 'Miesiąc', 'Kwota', 'Status']
        if header != expected_columns:
            create_emergency_backup(BUDGET_PATH)
            with open(BUDGET_PATH, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['ID', 'Rok', 'Miesiąc', 'Kwota', 'Status'])

def load_budgets():
    with open(BUDGET_PATH, 'r', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        next(reader)
        return list(reader)

def add_budget(id, year, month, amount):
    with open(BUDGET_PATH, 'a', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([id, year, month, amount])

def write_all_budgets(all_rows):
    all_rows.sort(key=lambda row: (int(row[1]), int(row[2])))
    with open(BUDGET_PATH, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['ID', 'Rok', 'Miesiąc', 'Kwota', 'Status'])
        writer.writerows(all_rows)

def resolve_export_path(output_arg, file_format='csv'):
    if os.sep in output_arg or '/' in output_arg:
        path = output_arg
        folder = os.path.dirname(path)
        if folder:
            os.makedirs(folder, exist_ok=True)
    else:
        os.makedirs(EXPORTS_DIR, exist_ok=True)
        path = os.path.join(EXPORTS_DIR, output_arg)
    expected_extension = f".{file_format}"
    if not path.endswith(expected_extension):
        path += expected_extension
    return path

def export_to_csv(data, file_path):
    with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerows(data)

def export_to_json(data, file_path, metadata=None):
    export_data = {
        'metadata': metadata if metadata else {},
        'data': data,
    }
    with open(file_path, 'w', newline='', encoding='utf-8') as jsonfile:
        json.dump(export_data, jsonfile, ensure_ascii=False, indent=4)

def export_to_excel(headers, data, file_path, metadata=None, sheet_name='Dane'):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    for row in data:
        ws.append(row)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    ws.freeze_panes = 'A2'
    if metadata:
        ws_meta = wb.create_sheet("Informacje")
        ws_meta.append(['Wyeksportowano:', metadata.get('exported_at', '')])
        ws_meta.append(['Źródło:', metadata.get('source', '')])
        ws_meta.append(['Liczba rekordów:', metadata.get('count', '')])
        if metadata.get('filters'):
            ws_meta.append(['Filtry:', ''])
            for key, value in metadata['filters'].items():
                ws_meta.append(['', f"{key}: {value}"])
    wb.save(file_path)